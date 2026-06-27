"""
Lendo uma planilha Excel com openpyxl.
Instale com:
pip install openpyxl
"""

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Instale a biblioteca com: pip install openpyxl")
else:
    arquivo = "vendas.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"
    ws.append(["produto", "quantidade"])
    ws.append(["X-Burguer", 12])
    ws.append(["Suco", 8])
    wb.save(arquivo)

    planilha = load_workbook(arquivo)
    aba = planilha["Vendas"]

    for produto, quantidade in aba.iter_rows(min_row=2, values_only=True):
        print(f"{produto}: {quantidade}")

